import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import pandas as pd
from urllib.parse import unquote

# Función para modificar el enlace solo en archivos específicos
def modify_link(original_link):
    if "sharepoint.com" in original_link:
        domain_end = original_link.find(".com") + 4
        if original_link.endswith(".xlsx"):
            new_link = original_link[:domain_end] + "/:x:/r" + original_link[domain_end:] + "?csf=1&web=1"
        elif original_link.endswith(".docx") or original_link.endswith(".doc"):
            new_link = original_link[:domain_end] + "/:w:/r" + original_link[domain_end:] + "?csf=1&web=1"
        elif original_link.endswith(".pptx") or original_link.endswith(".ppt"):
            new_link = original_link[:domain_end] + "/:p:/r" + original_link[domain_end:] + "?csf=1&web=1"
        else:
            new_link = original_link
        return new_link
    return original_link

# Función para crear una clave única basada en la ruta desde el archivo Excel
def generar_clave_unica(nombre, ruta, base_ruta):
    # Eliminar la parte de la ruta antes de base_ruta
    start_pos = ruta.find(base_ruta)
    ruta_relevante = ruta[start_pos:] if start_pos != -1 else ruta
    return f"{str(ruta_relevante).strip()}/{str(nombre).strip()}"

# Función para escribir el árbol en Excel, manteniendo el enlace original para carpetas
def write_tree_to_excel(tree, ws, hyperlink_dict, base_ruta, level=1, row=2, ruta_acumulada=""):
    folder_fill = PatternFill(start_color="FFE9A2", end_color="F8D775", fill_type="solid")
    link_fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Si `tree` es una lista, itera sobre cada elemento como si fueran nodos raíz
    if isinstance(tree, list):
        for item in tree:
            row = write_tree_to_excel(item, ws, hyperlink_dict, base_ruta, level, row, ruta_acumulada)
        return row

    if isinstance(tree, dict):
        ws.cell(row=row, column=level).value = tree['name']
        clave = generar_clave_unica(tree['name'], ruta_acumulada, base_ruta)
        # print("CLAVE EN WRITE_TREE",clave)
        
        # Verificar si la clave existe en el diccionario de enlaces
        if clave in hyperlink_dict:
            original_link = hyperlink_dict[clave]
            ws.cell(row=row, column=level).hyperlink = original_link
            ws.cell(row=row, column=level).font = Font(underline='single')
        
        ws.cell(row=row, column=level).fill = folder_fill
        ws.cell(row=row, column=level).border = border
        start_row = row
        row += 1
        
        # Llamada recursiva para cada hijo, asegurando que los archivos aparecen antes que las carpetas
        for child in tree.get('children', []):
            row = write_tree_to_excel(child, ws, hyperlink_dict, base_ruta, level + 1, row, ruta_acumulada + "/" + tree['name'])
        
        end_row = row - 1
        if end_row > start_row:
            ws.merge_cells(start_row=start_row, start_column=level, end_row=end_row, end_column=level)
            ws.cell(row=start_row, column=level).alignment = Alignment(horizontal='center', vertical='center')
            for r in range(start_row, end_row + 1):
                ws.cell(row=r, column=level).border = border

    else:
        ws.cell(row=row, column=level).value = tree
        clave = generar_clave_unica(tree, ruta_acumulada, base_ruta)

        # Solo agregar enlaces si es un archivo (tiene extensión)
        if clave in hyperlink_dict and '.' in str(tree):
            original_link = hyperlink_dict[clave]
            modified_link = modify_link(original_link)
            
            ws.cell(row=row, column=level + 1).value = "Link Online"
            ws.cell(row=row, column=level + 1).hyperlink = modified_link
            ws.cell(row=row, column=level + 1).fill = link_fill
            ws.cell(row=row, column=level + 1).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=level + 1).font = Font(underline='single')
            ws.cell(row=row, column=level + 1).border = border

            ws.cell(row=row, column=level + 2).value = "Link de Descarga"
            ws.cell(row=row, column=level + 2).hyperlink = original_link
            ws.cell(row=row, column=level + 2).fill = link_fill
            ws.cell(row=row, column=level + 2).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=level + 2).font = Font(underline='single')
            ws.cell(row=row, column=level + 2).border = border

        row += 1
        
    return row


# Modificación de la función para construir el árbol sin la carpeta principal como raíz
def build_tree_from_paths(dataframe, folder_column="Ruta de acceso", name_column="Nombre"):

    # Extraer la primera carpeta de la primera fila después de ordenar
    main_folder = dataframe.iloc[0][folder_column].strip().split('/')[-1]

    print(main_folder)

    # Diccionario para almacenar nodos de cada parte del árbol
    path_dict = {}

    # Construcción del árbol a partir de los paths en el dataframe
    for _, row in dataframe.iterrows():
        path_parts = row[folder_column].split('/')
        
        # Identificar la posición de la "carpeta principal" y omitirla
        if main_folder in path_parts:
            path_parts = path_parts[path_parts.index(main_folder) + 1:]
        
        current_level = path_dict

        # Recorrer o crear nodos para cada parte del path
        for part in path_parts:
            if part not in current_level or current_level[part] is None:
                current_level[part] = {}
            current_level = current_level[part]

        # Convertir el nombre del elemento a cadena para verificar si es un archivo
        item_name = str(row[name_column])
        if item_name not in current_level:
            # Marcar los archivos como None para distinguirlos de carpetas
            current_level[item_name] = None if '.' in item_name else {}

    # Función recursiva para convertir el diccionario en una estructura de árbol de listas
    def dict_to_tree(node):
        tree_node = []
        files = []
        folders = []
        for key, value in node.items():
            # Agrupar archivos y carpetas por separado
            if isinstance(value, dict):  # Carpeta
                folders.append({'name': key, 'children': dict_to_tree(value)})
            else:  # Archivo
                files.append(key)
        tree_node.extend(files + folders)  # Archivos primero, luego carpetas
        return tree_node

    # Convertir path_dict en una estructura de árbol sin la carpeta principal
    tree = dict_to_tree(path_dict)
    return tree


# Modificación de la función para cargar hipervínculos desde el archivo Excel sin incluir la carpeta principal en la clave
def load_hyperlinks_from_excel(file_path, base_ruta):
    hyperlink_wb = load_workbook(file_path)
    hyperlink_sheet = hyperlink_wb.active
    hyperlink_dict = {}

    # Extraer la última carpeta de base_ruta como "carpeta principal"
    main_folder = base_ruta.strip().split('/')[-1]

    for row in hyperlink_sheet.iter_rows(min_row=2, max_row=hyperlink_sheet.max_row, values_only=False):
        cell = row[0]
        ruta = row[5].value if len(row) > 5 else ""

        # Eliminar la carpeta principal de la ruta para que coincida con la estructura del árbol
        if main_folder in ruta:
            ruta = ruta.split(main_folder, 1)[-1]  # Deja solo la parte después de la carpeta principal

        if cell.hyperlink:
            clave = generar_clave_unica(cell.value, ruta, base_ruta)
            # print("clave en hyperlinks",clave)
            hyperlink_dict[clave] = cell.hyperlink.target

    return hyperlink_dict


# Modificación principal para escribir cada subcarpeta en una hoja distinta de Excel
def write_subfolders_to_sheets(tree_structure, workbook, hyperlink_dict, base_ruta):
    # Iterar sobre cada subcarpeta en la raíz del árbol
    for subfolder in tree_structure:
        # Verificar que `subfolder` es un diccionario con una clave `name` y `children`
        if isinstance(subfolder, dict) and 'name' in subfolder and 'children' in subfolder:
            # Crear una nueva hoja con el nombre de la subcarpeta
            ws = workbook.create_sheet(title=subfolder['name'][:31])  # Limita el nombre de la hoja a 31 caracteres
            ws.append(['', '', ''])
            
            # Llamar a `write_tree_to_excel` para escribir el contenido de la subcarpeta en la hoja
            write_tree_to_excel(subfolder, ws, hyperlink_dict, base_ruta)

def cambiar_nombres_carpetas(excel_file_path):

    # Cargar el archivo Excel utilizando openpyxl
    workbook = load_workbook(excel_file_path)

    # Seleccionar la primera hoja
    sheet = workbook.active

    # Función para verificar si un nombre tiene extensión de archivo
    def has_extension(name):
        return '.' in str(name).split('/')[-1]

    # Procesar cada celda en la columna "Nombre"
    for row in range(2, sheet.max_row + 1):  # Asumiendo que hay un encabezado en la primera fila
        cell = sheet[f'A{row}']  # Columna A corresponde a "Nombre"
        
        if cell.hyperlink and not has_extension(cell.value):
            # Extraer y decodificar la URL del hipervínculo
            url = unquote(cell.hyperlink.target)
            
            # Obtener la cadena anterior a "&View={"
            main_part = url.split('&View={')[0]
            
            # Dividir por "/" y obtener la última parte
            last_part = main_part.split('/')[-1]
            
            # Actualizar el valor en la celda conservando el hipervínculo
            cell.value = last_part
            cell.hyperlink = url  # Mantener el mismo hipervínculo

    # Guardar los cambios en el mismo archivo Excel
    workbook.save(excel_file_path)


# Ajustes al código principal para cargar el Excel y construir el árbol sin la carpeta principal
excel_file_path = 'query Proyecto comercial.xlsx'

cambiar_nombres_carpetas(excel_file_path)

excel_data = pd.read_excel(excel_file_path)



# Determinar la base de ruta a partir de la primera fila del Excel
excel_data = excel_data.sort_values(by='Ruta de acceso').reset_index(drop=True)

# Determinar la base de ruta a partir de la primera fila del Excel
base_ruta = excel_data.iloc[0]['Ruta de acceso']

# Cargar hipervínculos y construir el árbol desde la columna en el Excel
hyperlink_dict = load_hyperlinks_from_excel(excel_file_path, base_ruta)
tree_structure = build_tree_from_paths(excel_data)

# Crear el archivo de Excel y escribir cada subcarpeta en una hoja separada
wb = Workbook()
wb.remove(wb.active)  # Eliminar la hoja predeterminada vacía
write_subfolders_to_sheets(tree_structure, wb, hyperlink_dict, base_ruta)

# Guardar el archivo
wb.save('MAPEO_'+excel_file_path)
