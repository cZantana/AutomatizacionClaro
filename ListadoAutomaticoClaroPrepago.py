import os
import re
import fitz  # PyMuPDF para extraer texto de PDFs
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import requests
import json

import unicodedata

columnasReferencia = {
    "A":0,
    "B":1,
    "C":2,
    "D":3,
    "E":4,
    "F":5,
    "G":6,
}

# Definir un nombre fijo para el archivo temporal
TEMP_PDF_NAME = "archivo_temporal.pdf" # Archivo pdf temporal que se descarga automaticamente
EXCEL_GENERADO = "intentoListado Claro Prepago Febrero 2025.xlsx" #Nombre del archivo final generado

NombreListadoReferencia = "Listado_Referencia_Febrero_2025.xlsx" #Listado que da claro como inventario de los archivos
NombreMapaArchivosADescargar = 'MAPA_DE_ARCHIVOS_A_DESCARGAR.xlsx' #mapeo manual de los archivos que se van a descargar automaticamente

ColumnaComparar="E" #Columna con la cual se comparara la informacion dentro del archivo
ColumnaConectados="F" #Columna de la cual se obtendra el link de conectados
ColumnaCategoria = "C"  # Nueva columna para la categoría

NombreHojaReferencia = "MOVIL 2025" #Nombre de la hoja en la lista que da de inventario claro




# Expresión regular para encontrar "PCAM" o "PTAR" seguido de un número de 4 dígitos y opcionalmente " F" o "_F"
# pattern = re.compile(r'\b(PCAM|PTAR) \d{4}(?: F|_F|\b)', re.IGNORECASE)
asunto_pattern = re.compile(r'asunto\s*[:]?\s*((?:.|\n)*?)(?=fecha|proceso|$)', re.IGNORECASE)
fecha_pattern = re.compile(r'fecha\s*[:]?\s*(.*)', re.IGNORECASE | re.MULTILINE)

def remove_accents(input_str):
    nfkd_form = unicodedata.normalize('NFKD', input_str)
    return ''.join([char for char in nfkd_form if not unicodedata.combining(char)])


# Función para verificar si el nombre corresponde a un archivo PDF
def is_pdf(name):
    return bool(re.search(r'\.pdf$', name, re.IGNORECASE))

# Función para descargar un archivo PDF desde un enlace
def download_file(directory, link):
    """ Descarga un archivo desde un enlace y lo guarda con un nombre fijo """
    temp_pdf_path = os.path.join(directory, TEMP_PDF_NAME)

    try:
        with open("Cookies.JSON", "r") as f:
            cookies = json.load(f)
    except FileNotFoundError:
        print("No se encontró el archivo de cookies.")
        return None

    response = requests.get(link, cookies=cookies)
    if response.status_code == 200:
        with open(temp_pdf_path, 'wb') as f:
            f.write(response.content)
        print("Descarga completada.")
        return temp_pdf_path
    else:
        print(f"Error en la descarga: {response.status_code}")
        return None

# Función para extraer los primeros 10 renglones de un PDF
def extract_text_from_pdf(pdf_path):
    """ Extrae los primeros 20 renglones de un PDF y busca la cadena de interés """
    try:
        doc = fitz.open(pdf_path)
        text = []
        for page in doc:
            lines = page.get_text("text").split("\n")
            text.extend(lines)
            if len(text) >= 32:
                break
        extracted_text = "\n".join(text[:32])
        print(extracted_text)

        # Buscar la cadena de interés con regex
        asunto_match = asunto_pattern.search(extracted_text)
        fecha_match = fecha_pattern.search(extracted_text)

        # Validación antes de acceder a group(1)
        print("DESDE AQUI--------------------------------------------------------------------------")
        if asunto_match:
            print(f"ASUNTO: {asunto_match.group(1).replace(',', '').replace('\n', ' ').strip().rstrip('.')}")
        else:
            print("ASUNTO NO ENCONTRADO")

        if fecha_match:
            print(f"FECHA: {fecha_match.group(1).replace(',', '').strip().rstrip('.')}")
        else:
            print("FECHA NO ENCONTRADA")

        return {
            "asunto": asunto_match.group(1).replace(',', '').replace('\n', ' ').strip().rstrip('.').replace("  "," ") if asunto_match else "",
            "fecha": fecha_match.group(1).replace(',', '').strip().rstrip('.') if fecha_match else ""
        }

    except Exception as e:
        print(f"Error al extraer texto: {e}")
        return {"asunto": "", "fecha": ""}


    except Exception as e:
        print(f"Error al extraer texto: {e}")
        return { "asunto": "", "fecha": ""}

# Función para leer el archivo de referencias y almacenar sus datos
def load_reference_data(excel_path):
    """Carga los datos de la hoja LISTADO_CLIENTE_HOGAR y almacena las columnas B y D en un diccionario."""
    wb = load_workbook(excel_path)
    sheet = wb[NombreHojaReferencia]
    
    reference_data = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[columnasReferencia[ColumnaComparar]]:  # Verificar que la columna de comparación no esté vacía
            reference_data[row[columnasReferencia[ColumnaComparar]]] = (
                row[columnasReferencia[ColumnaConectados]],  # Valor de la columna de conectados
                row[columnasReferencia[ColumnaCategoria]]   # Valor de la columna de categoría
            )
    return reference_data

# Crear un nuevo archivo Excel cada vez que se ejecuta el script
def create_new_excel(file_name):
    """Crea un nuevo archivo Excel con los encabezados definidos."""
    wb = Workbook()
    sheet = wb.active
    sheet.append(["Categoría","Nombre_Archivo", "Politíca", "Link_de_descarga", "Link_Conectados", "Vigencia"])  # Encabezados
    wb.save(file_name)
    print(f"Se ha creado un nuevo archivo: {file_name}")

# Escribir datos en el archivo Excel con formato de botones
def write_to_excel(file_name, data, highlight):
    """Agrega datos al archivo Excel generado con formato de botones en enlaces."""
    wb = load_workbook(file_name)
    sheet = wb.active
    
    row = sheet.max_row + 1
    sheet.append(data)

    # Formato de celda para botones
    link_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    if highlight == False:
        # Configuración de botones para enlaces
        # sheet.cell(row=row, column=4, value="URL Documento Base").hyperlink = data[3]
        sheet.cell(row=row, column=4).fill = link_fill
        sheet.cell(row=row, column=4).alignment = Alignment(horizontal='center')
        sheet.cell(row=row, column=4).font = Font(underline='single', color='0000FF')
        sheet.cell(row=row, column=4).border = border
        
        sheet.cell(row=row, column=5).hyperlink = data[4]
        sheet.cell(row=row, column=5).fill = link_fill
        sheet.cell(row=row, column=5).alignment = Alignment(horizontal='center')
        sheet.cell(row=row, column=5).font = Font(underline='single', color='0000FF')
        sheet.cell(row=row, column=5).border = border
    
    else:

        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        sheet.cell(row=row, column=3).fill = red_fill
        sheet.cell(row=row, column=2).fill = red_fill
        
        # sheet.cell(row=row, column=4, value="URL Documento Base").hyperlink = data[3]
        sheet.cell(row=row, column=4).fill = link_fill
        sheet.cell(row=row, column=4).alignment = Alignment(horizontal='center')
        sheet.cell(row=row, column=4).font = Font(underline='single', color='0000FF')
        sheet.cell(row=row, column=4).border = border

    wb.save(file_name)

# Cargar los datos de referencia
reference_file = NombreListadoReferencia
reference_data = load_reference_data(reference_file)

# Crear un nuevo archivo Excel antes de iniciar el procesamiento
create_new_excel(EXCEL_GENERADO)

# Procesar archivos PDF
file_path = NombreMapaArchivosADescargar
wb = load_workbook(file_path)
sheet = wb.active
base_directory = os.path.abspath(os.path.dirname(file_path))
os.makedirs(base_directory, exist_ok=True)

for col in range(1, sheet.max_column + 1):
    if all(sheet.cell(row=row, column=col).value is None for row in range(1, sheet.max_row + 1)):
        break  # Termina si la columna está completamente vacía

    for row in range(1, sheet.max_row + 1):
        cell_value = sheet.cell(row=row, column=col).value
        if cell_value is None:
            continue  # Saltar celdas vacías
        
        cell_value = str(cell_value)

        if is_pdf(cell_value):
            link_column_index = col + 2
            link = sheet.cell(row=row, column=link_column_index).hyperlink.target if sheet.cell(row=row, column=link_column_index).hyperlink else None
            
            if link:
                pdf_path = download_file(base_directory, link)
                if pdf_path:
                    extracted_data = extract_text_from_pdf(pdf_path)
                    if extracted_data:
                        for name, (ref_link, category) in reference_data.items():
                            if remove_accents(extracted_data["asunto"].lower().replace(" ","").replace(",", "")) in remove_accents(name.strip().lower().replace(" ","").replace(",", "")):
                                write_to_excel(EXCEL_GENERADO, [category,cell_value,extracted_data["asunto"], link, ref_link, extracted_data["fecha"]], highlight=False)
                                break
                        else:
                            print(f"No se encontró coincidencia para el código: {extracted_data["asunto"]} en el archivo de referencias.")
                            write_to_excel(EXCEL_GENERADO, ["",cell_value, extracted_data["asunto"], link, "","",""], highlight=True)
                    else:
                        print(f"No se encontró un asunto en el archivo PDF: {cell_value}")

print("Proceso completado.")
