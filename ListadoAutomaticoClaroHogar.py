import os
import re
import fitz  # PyMuPDF para extraer texto de PDFs
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import requests
import json

columnasReferencia = {
    "A":0,
    "B":1,
    "C":2,
    "D":3,
    "E":4,
    "F":5,
    "G":6,
}

# Definir un nombre fijo para el archivo temporal
TEMP_PDF_NAME = "archivo_temporal.pdf"
EXCEL_GENERADO = "intentoListado Claro Hogar Febrero 2025.xlsx"

NombreListadoReferencia = "Listado_Referencia_Febrero_2025.xlsx"
NombreMapaArchivosADescargar = 'MAPA_DE_ARCHIVOS_A_DESCARGAR.xlsx'

ColumnaComparar="B" #Columna donde esta el nombre del archivo, para verificar que sea igual al titulo por dentro del pdf
ColumnaConectados="D" #Columna con los links de conectados
ColumnaCategoria="A" #Columna donde esa la categoria

NombreHojaReferencia = "HOGAR 2025" #Nombre de la hoja dentro del excel




# Expresión regular para encontrar "PCAM" o "PTAR" seguido de un número de 4 dígitos y opcionalmente " F" o "_F"
pattern = re.compile(r'\b(PCAM|PTAR) \d{4}(?: F|_F|\b)', re.IGNORECASE)
vigencia_pattern = re.compile(r'vigencia:\s*(.*)', re.IGNORECASE)
version_pattern = re.compile(r'versión:\s*(.*)', re.IGNORECASE)

# Función para verificar si el nombre corresponde a un archivo PDF
def is_pdf(name):
    return bool(re.search(r'\.pdf$', name, re.IGNORECASE))

# Función para descargar un archivo PDF desde un enlace
def download_file(directory, link):
    """ Descarga un archivo desde un enlace y lo guarda con un nombre fijo """
    temp_pdf_path = os.path.join(directory, TEMP_PDF_NAME)

    try:
        with open("Cookies.JSON", "r") as f:
            cookies = json.load(f)
    except FileNotFoundError:
        print("No se encontró el archivo de cookies.")
        return None

    response = requests.get(link, cookies=cookies)
    if response.status_code == 200:
        with open(temp_pdf_path, 'wb') as f:
            f.write(response.content)
        print("Descarga completada.")
        return temp_pdf_path
    else:
        print(f"Error en la descarga: {response.status_code}")
        return None

# Función para extraer los primeros 10 renglones de un PDF
def extract_text_from_pdf(pdf_path):
    """ Extrae los primeros 10 renglones de un PDF y busca la cadena de interés """
    try:
        doc = fitz.open(pdf_path)
        text = []
        for page in doc:
            lines = page.get_text("text").split("\n")  # Dividir por líneas
            text.extend(lines)  # Agregar líneas al texto total
            if len(text) >= 10:  # Si ya tenemos 10 líneas, terminamos
                break
        extracted_text = "\n".join(text[:10])
        print(extracted_text)

        # Buscar la cadena de interés con regex
        match = pattern.search(extracted_text)
        vigencia_match = vigencia_pattern.search(extracted_text)
        version_match = version_pattern.search(extracted_text)
        
        return {
            "codigo": match.group() if match else None,
            "vigencia": vigencia_match.group(1) if vigencia_match else "",
            "version": version_match.group(1) if version_match else ""
        }

    except Exception as e:
        print(f"Error al extraer texto: {e}")
        return {"codigo": None, "vigencia": "", "version": ""}

# Función para leer el archivo de referencias y almacenar sus datos
def load_reference_data(excel_path):
    """Carga los datos de la hoja LISTADO_CLIENTE_HOGAR y almacena las columnas B y D en un diccionario."""
    wb = load_workbook(excel_path)
    sheet = wb[NombreHojaReferencia]
    
    reference_data = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[columnasReferencia[ColumnaComparar]]:  # Verificar que la columna de comparación no esté vacía
            reference_data[row[columnasReferencia[ColumnaComparar]]] = (
                row[columnasReferencia[ColumnaConectados]],  # Valor de la columna de conectados
                row[columnasReferencia[ColumnaCategoria]]   # Valor de la columna de categoría
            )
    return reference_data

# Crear un nuevo archivo Excel cada vez que se ejecuta el script
def create_new_excel(file_name):
    """Crea un nuevo archivo Excel con los encabezados definidos."""
    wb = Workbook()
    sheet = wb.active
    sheet.append(["Categoría","Nombre Archivo", "Codigo Archivo", "Link de descarga", "Link Conectados", "Vigencia", "Versión"])  # Encabezados
    wb.save(file_name)
    print(f"Se ha creado un nuevo archivo: {file_name}")

# Escribir datos en el archivo Excel con formato de botones
def write_to_excel(file_name, data, highlight):
    """Agrega datos al archivo Excel generado con formato de botones en enlaces."""
    wb = load_workbook(file_name)
    sheet = wb.active
    
    row = sheet.max_row + 1
    sheet.append(data)
    
    if highlight == False:

        # Formato de celda para botones
        link_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Configuración de botones para enlaces
        # sheet.cell(row=row, column=4, value="URL Documento Base").hyperlink = data[3]
        sheet.cell(row=row, column=4).fill = link_fill
        sheet.cell(row=row, column=4).alignment = Alignment(horizontal='center')
        sheet.cell(row=row, column=4).font = Font(underline='single', color='0000FF')
        sheet.cell(row=row, column=4).border = border
        
        sheet.cell(row=row, column=5).hyperlink = data[4]
        sheet.cell(row=row, column=5).fill = link_fill
        sheet.cell(row=row, column=5).alignment = Alignment(horizontal='center')
        sheet.cell(row=row, column=5).font = Font(underline='single', color='0000FF')
        sheet.cell(row=row, column=5).border = border
    
    else:

        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        sheet.cell(row=row, column=3).fill = red_fill
        sheet.cell(row=row, column=2).fill = red_fill


    wb.save(file_name)

# Cargar los datos de referencia
reference_file = NombreListadoReferencia
reference_data = load_reference_data(reference_file)

# Crear un nuevo archivo Excel antes de iniciar el procesamiento
create_new_excel(EXCEL_GENERADO)

# Procesar archivos PDF
file_path = NombreMapaArchivosADescargar
wb = load_workbook(file_path)
sheet = wb.active
base_directory = os.path.abspath(os.path.dirname(file_path))
os.makedirs(base_directory, exist_ok=True)

for col in range(1, sheet.max_column + 1):
    if all(sheet.cell(row=row, column=col).value is None for row in range(1, sheet.max_row + 1)):
        break  # Termina si la columna está completamente vacía

    for row in range(1, sheet.max_row + 1):
        cell_value = sheet.cell(row=row, column=col).value
        if cell_value is None:
            continue  # Saltar celdas vacías
        
        cell_value = str(cell_value)

        if is_pdf(cell_value):
            link_column_index = col + 2
            link = sheet.cell(row=row, column=link_column_index).hyperlink.target if sheet.cell(row=row, column=link_column_index).hyperlink else None
            
            if link:
                pdf_path = download_file(base_directory, link)
                if pdf_path:
                    extracted_data = extract_text_from_pdf(pdf_path)
                    if extracted_data:
                        for name, (ref_link, category) in reference_data.items():
                            if extracted_data["codigo"] in name:
                                write_to_excel(EXCEL_GENERADO, [category, cell_value, extracted_data["codigo"].replace("_F","F"), link, ref_link, extracted_data["vigencia"], extracted_data["version"]], highlight=False)
                                break
                        else:
                            print(f"No se encontró coincidencia para el código: {extracted_data} en el archivo de referencias.")
                            write_to_excel(EXCEL_GENERADO, ["","Código no coincidente en nombre de archivo", extracted_data["codigo"].replace('_F', 'F'), "", "","",""], highlight=True)
                    else:
                        print(f"No se encontró un código en el archivo PDF: {cell_value}")

print("Proceso completado.")
